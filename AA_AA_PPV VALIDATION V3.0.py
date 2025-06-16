import os
import glob
import time
import pytz
import json
import requests
import pandas as pd
from datetime import datetime, timedelta
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path


def download_latest_csv(url, download_path):
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    csv_links = [a['href'] for a in soup.find_all('a', href=True) if a['href'].endswith('.csv')]

    if not csv_links:
        raise ValueError("No CSV files found at the provided URL.")
    
    latest_csv_url = csv_links[-1]
    full_csv_url = urljoin(url, latest_csv_url)
    csv_response = requests.get(full_csv_url)
    csv_response.raise_for_status()
    
    csv_file_name = os.path.basename(full_csv_url)
    csv_file_path = os.path.join(download_path, csv_file_name)
    
    with open(csv_file_path, 'wb') as file:
        file.write(csv_response.content)

    print(f"Downloaded the latest CSV file: {csv_file_name} to {csv_file_path}")

def get_csv_file():
    script_directory = os.path.dirname(os.path.abspath(__file__))
    csv_files = glob.glob(os.path.join(script_directory, '*.csv'))
    
    if len(csv_files) != 1:
        raise ValueError(f"Expected exactly one CSV file in the folder, but found {len(csv_files)}.")
    
    return os.path.basename(csv_files[0])

def parse_custom_time(time_str):
    time_parts = time_str.split(':')
    hours = int(time_parts[0])
    minutes = int(time_parts[1][:-1])
    is_pm = time_parts[1][-1].lower() == 'p'
    
    if is_pm and hours != 12:
        hours += 12
    elif not is_pm and hours == 12:
        hours = 0
        
    return hours, minutes

def convert_to_utc(event_date, event_time):
    hours, minutes = parse_custom_time(event_time)
    current_year = datetime.now().year
    event_datetime_str = f"{event_date} {current_year} {hours:02d}:{minutes:02d}"
    event_datetime = datetime.strptime(event_datetime_str, '%A %B %d %Y %H:%M')
    
    est_timezone = pytz.timezone('US/Eastern')
    event_datetime_est = est_timezone.localize(event_datetime)
    event_datetime_utc = event_datetime_est.astimezone(pytz.utc)
    
    return event_datetime_utc

def convert_and_rename_excel(csv_path, event_name):
    df = pd.read_csv(csv_path, low_memory=False)
    current_name = os.path.basename(csv_path)
    file_number = current_name.split('_')[4].split('.')[0]
    new_name = str(Path.cwd() / download_path / f'IP PPV {event_name}_{file_number}.xlsx')

    
    with pd.ExcelWriter(new_name, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=os.path.splitext(current_name)[0])
    
    print("File converted and renamed successfully.")
    return new_name

def insert_datetime(sheet, column, datetime_value, row):
    sheet[f"{column}{row}"] = datetime_value

def color_code_cells(sheet, column, value_to_match, start_row):
    good_font_color = '006100'
    good_fill_color = 'C6EFCE'
    bad_font_color = '9C0006'
    bad_fill_color = 'FFC7CE'
    
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f"{column}{row}"]
        if cell.value == value_to_match:
            cell.font = Font(color=good_font_color)
            cell.fill = PatternFill(start_color=good_fill_color, end_color=good_fill_color, fill_type="solid")
        else:
            cell.font = Font(color=bad_font_color)
            cell.fill = PatternFill(start_color=bad_fill_color, end_color=bad_fill_color, fill_type="solid")

def color_code_prices(sheet, column, price_to_match, start_row):
    good_font_color = '006100'
    good_fill_color = 'C6EFCE'
    bad_font_color = '9C0006'
    bad_fill_color = 'FFC7CE'
    
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f"{column}{row}"]
        if cell.value == price_to_match:
            cell.font = Font(color=good_font_color)
            cell.fill = PatternFill(start_color=good_fill_color, end_color=good_fill_color, fill_type="solid")
        else:
            cell.font = Font(color=bad_font_color)
            cell.fill = PatternFill(start_color=bad_fill_color, end_color=bad_fill_color, fill_type="solid")

def color_code_dates(sheet, column, date_to_match, start_row):
    good_font_color = '006100'
    good_fill_color = 'C6EFCE'
    bad_font_color = '9C0006'
    bad_fill_color = 'FFC7CE'
    
    for row in range(start_row, sheet.max_row + 1):
        cell = sheet[f"{column}{row}"]
        if cell.value == date_to_match:
            cell.font = Font(color=good_font_color)
            cell.fill = PatternFill(start_color=good_fill_color, end_color=good_fill_color, fill_type="solid")
        else:
            cell.font = Font(color=bad_font_color)
            cell.fill = PatternFill(start_color=bad_fill_color, end_color=bad_fill_color, fill_type="solid")

def find_next_empty_row(sheet, column, start_row):
    for row in range(start_row, sheet.max_row + 1):
        if sheet[f"{column}{row}"].value is None:
            return row
    return sheet.max_row + 1

def compare_and_color_code(sheet, master_values):
    good_font_color = '006100'
    good_fill_color = 'C6EFCE'
    bad_font_color = '9C0006'
    bad_fill_color = 'FFC7CE'
    
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f"J{row}"]
        value = cell.value
        
        if value and value in master_values:
            cell.font = Font(color=good_font_color)
            cell.fill = PatternFill(start_color=good_fill_color, end_color=good_fill_color, fill_type="solid")
        else:
            cell.font = Font(color=bad_font_color)
            cell.fill = PatternFill(start_color=bad_fill_color, end_color=bad_fill_color, fill_type="solid")

def compare_and_color_code_sheets(sheet, start_row):
    good_font_color = '006100'
    good_fill_color = 'C6EFCE'
    bad_font_color = '9C0006'
    bad_fill_color = 'FFC7CE'
    
    max_row = sheet.max_row
    
    for row in range(start_row, max_row + 1):
        cell_a = sheet.cell(row=row, column=1).value
        cell_b = sheet.cell(row=row, column=2).value
        
        found_match = False
        
        for b_row in range(start_row, max_row + 1):
            if cell_a == sheet.cell(row=b_row, column=2).value:
                found_match = True
                break
        
        if found_match:
            font_color = good_font_color
            fill_color = good_fill_color
        else:
            font_color = bad_font_color
            fill_color = bad_fill_color
        
        font_a = Font(color=font_color)
        fill_a = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        sheet.cell(row=row, column=1).font = font_a
        sheet.cell(row=row, column=1).fill = fill_a
        
        font_b = Font(color=font_color)
        fill_b = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        sheet.cell(row=row, column=2).font = font_b
        sheet.cell(row=row, column=2).fill = fill_b

def fetch_filtered_tag_names(guid, billing_id):
    url = f"http://inspector.merlin.comcast.net:8080/offerObjects?guid={guid}"
    response = requests.get(url)
    data = response.json()
    filtered_tag_names = []

    for offer in data.get('offers', []):
        if offer.get('billingId') == billing_id:
            for avail in offer.get('availabilities', []):
                tag_name = avail.get('availabilityTagName', '')
                if tag_name.startswith("Corp:") and not tag_name.startswith("Corp:8069") and not tag_name.startswith("Corp:8045"):
                    filtered_tag_names.append(tag_name)
    
    return filtered_tag_names

def update_excel_with_tag_names(file_path, hd_data, sd_data, es_data):
    wb = load_workbook(file_path)
    if hd_data:
        hd_sheet = wb['HD Availabilities']
        for row_idx, tag_name in enumerate(hd_data, start=1):
            hd_sheet.cell(row=row_idx, column=2, value=tag_name)
        compare_and_color_code_sheets(hd_sheet, 1)

    if sd_data:
        sd_sheet = wb['SD Availabilities']
        for row_idx, tag_name in enumerate(sd_data, start=1):
            sd_sheet.cell(row=row_idx, column=2, value=tag_name)
        compare_and_color_code_sheets(sd_sheet, 1)
        
    if es_data:
        es_sheet = wb['ES Availabilities']
        for row_idx, tag_name in enumerate(es_data, start=1):
            es_sheet.cell(row=row_idx, column=2, value=tag_name)
        compare_and_color_code_sheets(es_sheet, 1)

    wb.save(file_path)

def fetch_html_content(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.text
    else:
        print(f"Failed to retrieve HTML content: {response.status_code}")
        return None

def parse_listing_ids(html_content, event_name):
    soup = BeautifulSoup(html_content, 'html.parser')
    hd_ids, sd_ids, es_ids = [], [], []
    hd_pids, sd_pids, es_pids = [], [], []
    hd_sids, sd_sids, es_sids = [], [], []
    hd_cids, sd_cids, es_cids = [], [], []

    hd_items = soup.find_all(lambda tag: tag.name == 'a' and 
                             tag.get('data-grid-under') == 'grid-under-504' and
                             any(keyword.lower() in tag.get_text(strip=True).lower() for keyword in event_name.split()))

    sd_items = soup.find_all(lambda tag: tag.name == 'a' and 
                             tag.get('data-grid-under') == 'grid-under-501' and
                             any(keyword.lower() in tag.get_text(strip=True).lower() for keyword in event_name.split()))
    
    es_items = soup.find_all(lambda tag: tag.name == 'a' and 
                             tag.get('data-grid-under') == 'grid-under-502' and
                             any(keyword.lower() in tag.get_text(strip=True).lower() for keyword in event_name.split()))

    for item in hd_items:
        listing_id = item.get('data-listingid')
        if listing_id:
            hd_ids.append(listing_id)
        program_id = item.get('data-merlinid')
        if program_id:
            hd_pids.append(program_id)
        station_id = item.get('data-stationid')
        if station_id:
            hd_sids.append(station_id)
        channel_id = item.get('data-channelid')
        if channel_id:
            hd_cids.append(channel_id)

    for item in sd_items:
        listing_id = item.get('data-listingid')
        if listing_id:
            sd_ids.append(listing_id)
        program_id = item.get('data-merlinid')
        if program_id:
            sd_pids.append(program_id)
        station_id = item.get('data-stationid')
        if station_id:
            sd_sids.append(station_id)
        channel_id = item.get('data-channelid')
        if channel_id:
            sd_cids.append(channel_id)
            
    for item in es_items:
        listing_id = item.get('data-listingid')
        if listing_id:
            es_ids.append(listing_id)
        program_id = item.get('data-merlinid')
        if program_id:
            es_pids.append(program_id)
        station_id = item.get('data-stationid')
        if station_id:
            es_sids.append(station_id)
        channel_id = item.get('data-channelid')
        if channel_id:
            es_cids.append(channel_id)

    return hd_ids, sd_ids, hd_pids, sd_pids, hd_cids, sd_cids, hd_sids, sd_sids, es_ids, es_pids, es_sids, es_cids

def get_media_guid(url, params, settlement_reference):
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = response.json()
        entries = data.get('entries', [])

        for entry in entries:
            offer_media_associations = entry.get('offerMediaAssociations', [])
            for association in offer_media_associations:
                if association.get('settlementReference') == settlement_reference:
                    media_guid = association.get('mediaId', {}).get('mediaGuid')
                    if media_guid:
                        return media_guid
    else:
        print(f"Failed to retrieve data: {response.status_code}")
    return None

def main():
    
    start_time = time.time()  # Start the timer

    url = "https://vcwarchive.g.comcast.net/vcwh_exports/ppv/"

    user_home = Path.home()
    download_path = user_home / "PPV_Validation_Outputs"
    download_path.mkdir(exist_ok=True)

    download_latest_csv(url, download_path)

    csv_file = get_csv_file()
    event_name = input("Enter the event name: ")
    output_filename = convert_and_rename_excel(csv_file, event_name)

    HD_value_to_match = input("Enter the HD value to match (leave blank to skip): ")
    SD_value_to_match = input("Enter the SD value to match (leave blank to skip): ")
    ES_value_to_match = input("Enter the ES value to match (leave blank to skip): ")
    HD_Price = input("Enter the HD Price (leave blank to skip): ")
    SD_Price = input("Enter the SD Price (leave blank to skip): ")
    ES_Price = input("Enter the ES Price (leave blank to skip): ")
    Event_broadcast_date = input("Enter the event broadcast date (e.g., Saturday June 15): ")
    Event_broadcast_time = input("Enter the event countdown time (e.g., 7:00p for PM or 7:00a for AM): ")
    HD_SID = '13503'
    SD_SID = '12162'
    ES_SID = '15006'

    event_datetime_combined = convert_to_utc(Event_broadcast_date, Event_broadcast_time)
    event_end = event_datetime_combined + timedelta(hours=12.983333333333333)
    date_to_match = event_datetime_combined.strftime('%m/%d/%Y %H.%M.%S')
    event_datetime_combined_iso = event_datetime_combined.strftime('%Y-%m-%dT%H:%MZ')

    print("Here is the time and date value in UTC:")
    print(event_datetime_combined.strftime('%m/%d/%Y %H.%M.%S'))
    print(event_end.strftime('%m/%d/%Y %H.%M.%S'))
    print(f"Formatted datetime for URL: {event_datetime_combined_iso}")

    url = f"http://inspector.merlin.comcast.net:8080/loadGrid?accountId=7876220869746444319&startDate={event_datetime_combined_iso}&clientProfile=XRE:X2&supportedCatalogs=TitleVI,CTV&freeToMe=off"

    original_wb = load_workbook(output_filename)
    
    if HD_value_to_match:
        HD_new_sheet = original_wb.create_sheet(title='HD - ' + HD_value_to_match)
        print("Created HD sheet in the workbook")
    if SD_value_to_match:
        SD_new_sheet = original_wb.create_sheet(title='SD - ' + SD_value_to_match)
        print("Created SD sheet in the workbook")
    if ES_value_to_match:
        ES_new_sheet = original_wb.create_sheet(title='ES - ' + ES_value_to_match)
        print("Created ES sheet in the workbook")

    df = pd.read_excel(output_filename)
    print("Read data from the Excel")

    if HD_value_to_match:
        HD_new_sheet.append(df.columns.tolist())
    if SD_value_to_match:
        SD_new_sheet.append(df.columns.tolist())
    if ES_value_to_match:
        ES_new_sheet.append(df.columns.tolist())

    for index, row in df.iterrows():
        if HD_value_to_match and row['Billing Event Id'] == HD_value_to_match and row['Source Id'] == int(HD_SID) and not str(row['Corp']).startswith(('8069', '8045')):
            HD_new_sheet.append(row.tolist())
        if SD_value_to_match and row['Billing Event Id'] == SD_value_to_match and row['Source Id'] == int(SD_SID) and not str(row['Corp']).startswith(('8069', '8045')):
            SD_new_sheet.append(row.tolist())
        if ES_value_to_match and row['Billing Event Id'] == ES_value_to_match and row['Source Id'] == int(ES_SID) and not str(row['Corp']).startswith(('8069', '8045')):
            ES_new_sheet.append(row.tolist())

    if HD_value_to_match:
        color_code_cells(HD_new_sheet, 'H', HD_value_to_match, 2)
        color_code_prices(HD_new_sheet, 'I', float(HD_Price), 2)
        color_code_dates(HD_new_sheet, 'K', date_to_match, 2)
    if SD_value_to_match:
        color_code_cells(SD_new_sheet, 'H', SD_value_to_match, 2)
        color_code_prices(SD_new_sheet, 'I', float(SD_Price), 2)
        color_code_dates(SD_new_sheet, 'K', date_to_match, 2)
    if ES_value_to_match:
        color_code_cells(ES_new_sheet, 'H', ES_value_to_match, 2)
        color_code_prices(ES_new_sheet, 'I', float(ES_Price), 2)
        color_code_dates(ES_new_sheet, 'K', date_to_match, 2)

    if HD_value_to_match:
        hd_next_row = find_next_empty_row(HD_new_sheet, 'H', 2) + 1
        HD_new_sheet[f"H{hd_next_row}"] = HD_value_to_match
        HD_new_sheet[f"H{hd_next_row}"].font = Font(color='006100')
        HD_new_sheet[f"H{hd_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        hd_price_next_row = find_next_empty_row(HD_new_sheet, 'I', 2) + 1
        HD_new_sheet[f"I{hd_price_next_row}"] = float(HD_Price)
        HD_new_sheet[f"I{hd_price_next_row}"].font = Font(color='006100')
        HD_new_sheet[f"I{hd_price_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        hd_date_next_row = find_next_empty_row(HD_new_sheet, 'K', 2) + 1
        HD_new_sheet[f"K{hd_date_next_row}"] = date_to_match
        HD_new_sheet[f"K{hd_date_next_row}"].font = Font(color='006100')
        HD_new_sheet[f"K{hd_date_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

    if SD_value_to_match:
        sd_next_row = find_next_empty_row(SD_new_sheet, 'H', 2) + 1
        SD_new_sheet[f"H{sd_next_row}"] = SD_value_to_match
        SD_new_sheet[f"H{sd_next_row}"].font = Font(color='006100')
        SD_new_sheet[f"H{sd_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        sd_price_next_row = find_next_empty_row(SD_new_sheet, 'I', 2) + 1
        SD_new_sheet[f"I{sd_price_next_row}"] = float(SD_Price)
        SD_new_sheet[f"I{sd_price_next_row}"].font = Font(color='006100')
        SD_new_sheet[f"I{sd_price_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        sd_date_next_row = find_next_empty_row(SD_new_sheet, 'K', 2) + 1
        SD_new_sheet[f"K{sd_date_next_row}"] = date_to_match
        SD_new_sheet[f"K{sd_date_next_row}"].font = Font(color='006100')
        SD_new_sheet[f"K{sd_date_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

    if ES_value_to_match:
        es_next_row = find_next_empty_row(ES_new_sheet, 'H', 2) + 1
        ES_new_sheet[f"H{es_next_row}"] = ES_value_to_match
        ES_new_sheet[f"H{es_next_row}"].font = Font(color='006100')
        ES_new_sheet[f"H{es_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        es_price_next_row = find_next_empty_row(ES_new_sheet, 'I', 2) + 1
        ES_new_sheet[f"I{es_price_next_row}"] = float(ES_Price)
        ES_new_sheet[f"I{es_price_next_row}"].font = Font(color='006100')
        ES_new_sheet[f"I{es_price_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

        es_date_next_row = find_next_empty_row(ES_new_sheet, 'K', 2) + 1
        ES_new_sheet[f"K{es_date_next_row}"] = date_to_match
        ES_new_sheet[f"K{es_date_next_row}"].font = Font(color='006100')
        ES_new_sheet[f"K{es_date_next_row}"].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")

    master_corp_path = str(download_path / "MasterCorp.xlsx")
    master_corp_wb = load_workbook(master_corp_path)
    corp_sheet = master_corp_wb["Corp"]
    corp_values = [corp_sheet[f"A{i}"].value for i in range(1, 117)]

    if HD_value_to_match:
        hd_corp_next_row = find_next_empty_row(HD_new_sheet, 'J', 2) + 1
        for i, value in enumerate(corp_values):
            HD_new_sheet[f"J{hd_corp_next_row + i}"] = value

    if SD_value_to_match:
        sd_corp_next_row = find_next_empty_row(SD_new_sheet, 'J', 2) + 1
        for i, value in enumerate(corp_values):
            SD_new_sheet[f"J{sd_corp_next_row + i}"] = value

    if ES_value_to_match:
        es_corp_next_row = find_next_empty_row(ES_new_sheet, 'J', 2) + 1
        for i, value in enumerate(corp_values):
            ES_new_sheet[f"J{es_corp_next_row + i}"] = value

    print("Copied values from Corp sheet to HD, SD, and ES sheets one cell below the last entry in column J")

    def compare_and_color_code_both_sheets(hd_sheet, sd_sheet, es_sheet, master_values):
        good_font_color = '006100'
        good_fill_color = 'C6EFCE'
        bad_font_color = '9C0006'
        bad_fill_color = 'FFC7CE'

        def mark_cell(cell, condition):
            if condition:
                cell.font = Font(color=good_font_color)
                cell.fill = PatternFill(start_color=good_fill_color, end_color=good_fill_color, fill_type="solid")
            else:
                cell.font = Font(color=bad_font_color)
                cell.fill = PatternFill(start_color=bad_fill_color, end_color=bad_fill_color, fill_type="solid")

        if hd_sheet:
            for row in range(2, hd_sheet.max_row + 1):
                hd_cell = hd_sheet[f"J{row}"]
                if hd_cell.value is not None:
                    mark_cell(hd_cell, hd_cell.value in master_values)
            for i in range(len(master_values)):
                hd_copied_cell = hd_sheet[f"J{hd_corp_next_row + i}"]
                mark_cell(hd_copied_cell, True)

        if sd_sheet:
            for row in range(2, sd_sheet.max_row + 1):
                sd_cell = sd_sheet[f"J{row}"]
                if sd_cell.value is not None:
                    mark_cell(sd_cell, sd_cell.value in master_values)
            for i in range(len(master_values)):
                sd_copied_cell = sd_sheet[f"J{sd_corp_next_row + i}"]
                mark_cell(sd_copied_cell, True)

        if es_sheet:
            for row in range(2, es_sheet.max_row + 1):
                es_cell = es_sheet[f"J{row}"]
                if es_cell.value is not None:
                    mark_cell(es_cell, es_cell.value in master_values)
            for i in range(len(master_values)):
                es_copied_cell = es_sheet[f"J{es_corp_next_row + i}"]
                mark_cell(es_copied_cell, True)

    compare_and_color_code_both_sheets(
        HD_new_sheet if HD_value_to_match else None,
        SD_new_sheet if SD_value_to_match else None,
        ES_new_sheet if ES_value_to_match else None,
        corp_values
    )

    if HD_value_to_match:
        HD_availabilities_sheet = original_wb.create_sheet(title='HD Availabilities')
    if SD_value_to_match:
        SD_availabilities_sheet = original_wb.create_sheet(title='SD Availabilities')
    if ES_value_to_match:
        ES_availabilities_sheet = original_wb.create_sheet(title='ES Availabilities')
    
    corp_avail_sheet = master_corp_wb["Corp Availability"]
    for row in corp_avail_sheet.iter_rows():
        if HD_value_to_match:
            HD_availabilities_sheet.append([cell.value for cell in row])
        if SD_value_to_match:
            SD_availabilities_sheet.append([cell.value for cell in row])
        if ES_value_to_match:
            ES_availabilities_sheet.append([cell.value for cell in row])

    print("Copied data from 'Corp Availability' sheet to 'HD Availabilities', 'SD Availabilities' and 'ES Availabilities' sheets")

    original_wb.save(output_filename)
    print("Saved the changes to the original workbook\n\n")

    html_content = fetch_html_content(url)
    if html_content:
        HD_listing_id, SD_listing_id, HD_program_id, SD_program_id, HD_channel_id, SD_channel_id, HD_station_id, SD_station_id, ES_listing_id, ES_program_id, ES_channel_id, ES_station_id = parse_listing_ids(html_content, event_name)

        if HD_value_to_match and HD_listing_id:
            HD_url = "http://bo.prod.merlin.ccp.xcal.tv:9023/offerDataService/data/Offer"
            HD_params = {
                "schema": "2.34.0",
                "form": "cjson",
                "pretty": "true",
                "byOfferEntityAssociations.entityId": HD_listing_id
            }
            HD_settlement_reference = HD_value_to_match
            HD_Media_Guid = get_media_guid(HD_url, HD_params, HD_settlement_reference)

        if SD_value_to_match and SD_listing_id:
            SD_url = "http://bo.prod.merlin.ccp.xcal.tv:9023/offerDataService/data/Offer"
            SD_params = {
                "schema": "2.34.0",
                "form": "cjson",
                "pretty": "true",
                "byOfferEntityAssociations.entityId": SD_listing_id
            }
            SD_settlement_reference = SD_value_to_match
            SD_Media_Guid = get_media_guid(SD_url, SD_params, SD_settlement_reference)
            
        if ES_value_to_match and ES_listing_id:
            ES_url = "http://bo.prod.merlin.ccp.xcal.tv:9023/offerDataService/data/Offer"
            ES_params = {
                "schema": "2.34.0",
                "form": "cjson",
                "pretty": "true",
                "byOfferEntityAssociations.entityId": ES_listing_id
            }
            ES_settlement_reference = ES_value_to_match
            ES_Media_Guid = get_media_guid(ES_url, ES_params, ES_settlement_reference)
            
            print(f"---Details for ticket creation---\n")
            print(f"PPV Validations - {event_name}")
            print(f"{Event_broadcast_date}\n")
            print(f"Event time: {Event_broadcast_time}\n")
            if HD_value_to_match:
                print(f"HD Price: {HD_Price}")
            if SD_value_to_match:
                print(f"SD Price: {SD_Price}")
            if ES_value_to_match:
                print(f"ES Price: {ES_Price}\n\n")
            if HD_value_to_match:
                print(f"HD: {HD_value_to_match}")
                print(f"listing_id: {HD_listing_id}")
                print(f"Program HD ID: {HD_program_id}")
                print(f"Channel HD ID: {HD_channel_id}")
                print(f"Station HD ID: {HD_station_id}")
                print(f"HD_Media_Guid: {HD_Media_Guid}\n\n")
            if SD_value_to_match:
                print(f"SD: {SD_value_to_match}")
                print(f"listing_id: {SD_listing_id}")
                print(f"Program SD ID: {SD_program_id}")
                print(f"Channel SD ID: {SD_channel_id}")
                print(f"Station SD ID: {SD_station_id}")
                print(f"SD_Media_Guid: {SD_Media_Guid}\n")
            if ES_value_to_match:
                print(f"ES: {ES_value_to_match}")
                print(f"listing_id: {ES_listing_id}")
                print(f"Program ES ID: {ES_program_id}")
                print(f"Channel ES ID: {ES_channel_id}")
                print(f"Station ES ID: {ES_station_id}")
                print(f"ES_Media_Guid: {ES_Media_Guid}\n")

    hd_guid = HD_Media_Guid if HD_value_to_match else None
    sd_guid = SD_Media_Guid if SD_value_to_match else None
    es_guid = ES_Media_Guid if ES_value_to_match else None
    hd_billing_id = HD_value_to_match if HD_value_to_match else None
    sd_billing_id = SD_value_to_match if SD_value_to_match else None
    es_billing_id = ES_value_to_match if ES_value_to_match else None

    filtered_hd_names = fetch_filtered_tag_names(hd_guid, hd_billing_id) if hd_guid and hd_billing_id else []
    filtered_sd_names = fetch_filtered_tag_names(sd_guid, sd_billing_id) if sd_guid and sd_billing_id else []
    filtered_es_names = fetch_filtered_tag_names(es_guid, es_billing_id) if es_guid and es_billing_id else []

    update_excel_with_tag_names(output_filename, filtered_hd_names, filtered_sd_names, filtered_es_names)

    end_time = time.time()  # End the timer
    execution_time = end_time - start_time  # Calculate the total execution time
    print(f"PPV Validation for the asset {event_name} was completed in: {execution_time:.2f} seconds")

if __name__ == "__main__":
    main()
